import os
import subprocess
import csv
from codecarbon import EmissionsTracker
from datetime import datetime
import time
import pandas as pd
import shutil
import matplotlib.pyplot as plt
from fpdf import FPDF
from dotenv import load_dotenv
import seaborn as sns
from openpyxl import load_workbook

# Load environment variables
env_path = os.path.abspath(".env")
load_dotenv(dotenv_path=env_path, verbose=True, override=True)

SOURCE_DIRECTORY = os.path.dirname(env_path)
GREEN_REFINED_DIRECTORY = os.path.join(SOURCE_DIRECTORY, 'GreenCode')
RESULT_DIR = os.path.join(SOURCE_DIRECTORY, 'Result')

ExcludedFiles = ['server_emissions.py', 'GreenCodeRefiner.py', 'track_emissions.py', 'compare_emissions.py', 'GreenCode']

# Function to process emissions for a single file
def process_emissions_for_file(tracker, script_path, emissions_csv, file_type, result_dir, test_command):
    emissions_data = None
    duration = 0
    test_output = 'Unknown'
    script_name = os.path.basename(script_path)
    
    try:
        tracker.start()  # Start the emissions tracking
        
        if 'test' in script_name.lower():
            start_time = time.time()
            test_result = subprocess.run(test_command, capture_output=True, text=True, timeout=20)  # Run the test
            duration = time.time() - start_time
            test_output = 'Pass' if test_result.returncode == 0 else 'Fail'
        else:
            # This is a normal programming file (not a test)
            test_output = 'Not a test file'
            print(f"Skipping test execution for {script_name} as it is a normal programming file.")
    
    except subprocess.TimeoutExpired:
        test_output = 'Timeout'
        print(f"Test execution for {script_path} exceeded the timeout limit.")
    
    except Exception as e:
        test_output = 'Error'
        print(f"An error occurred for {script_path}: {e}")
    
    finally:
        emissions_data = tracker.stop()  # Stop the emissions tracking
    
    # If emissions data was collected, save it
    if emissions_data is not None:
        emissions_csv_default_path = 'emissions.csv'  # Default location for emissions.csv
        emissions_csv_target_path = os.path.join(result_dir, 'emissions.csv')  # Target location
        # Move the emissions.csv to the result directory
        if os.path.exists(emissions_csv_default_path):
            shutil.move(emissions_csv_default_path, emissions_csv_target_path)
        # Read the latest emissions data from the moved CSV
        if os.stat(emissions_csv_target_path).st_size != 0:
            emissions_data = pd.read_csv(emissions_csv_target_path).iloc[-1]
            data = [
                os.path.basename(script_path),
                file_type,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                emissions_data['emissions'] * 1000,  # Convert to gCO2eq
                duration,
                emissions_data['emissions_rate'] * 1000,  # Convert to gCO2eq/s
                emissions_data['cpu_power'],
                emissions_data['gpu_power'],
                emissions_data['ram_power'],
                emissions_data['cpu_energy'] * 1000,  # Convert to Wh
                emissions_data['gpu_energy'],
                emissions_data['ram_energy'] * 1000,  # Convert to Wh
                emissions_data['energy_consumed'] * 1000,  # Convert to Wh
                test_output
            ]
            with open(emissions_csv, 'a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(data)
                file.flush()
        else:
            print(f"No emissions data found for {script_path}")

# Function to process test execution for different file types
def process_files_by_type(base_dir, emissions_data_csv, result_dir, file_extension, excluded_files, tracker, test_command_generator):
    files = []
    for root, dirs, file_list in os.walk(base_dir):
        for script in file_list:
            if script.endswith(file_extension) and script not in excluded_files:
                files.append(os.path.join(root, script))
    
    for script_path in files:
        test_command = test_command_generator(script_path)
        process_emissions_for_file(
            tracker=tracker,
            script_path=script_path,
            emissions_csv=emissions_data_csv,
            file_type=file_extension,
            result_dir=result_dir,
            test_command=test_command
        )

# Generate test commands for each language
def get_python_test_command(script_path):
    return [os.getenv('PYTEST_PATH'), script_path]

def get_java_test_command(script_path):
    return [os.getenv('MAVEN_PATH'), '-Dtest=' + os.path.splitext(os.path.basename(script_path))[0] + 'Test', 'test']

def get_cpp_test_command(script_path):
    test_file_name = os.path.basename(script_path).replace('.cpp', '_test.cpp')
    test_file_path = os.path.join('test', test_file_name)
    compile_command = ['g++', '-o', 'test_output', test_file_path, '-lgtest', '-lgtest_main', '-pthread']
    run_command = ['./test_output']
    return compile_command + run_command

def get_cs_test_command(script_path):
    return [os.getenv('NUNIT_PATH'), 'test', os.path.splitext(os.path.basename(script_path))[0] + '.dll']

# Refactored process_folder function
def process_folder(base_dir, emissions_data_csv, result_dir, suffix):
    excluded_files = ['server_emissions.py', 'GreenCodeRefiner.py', 'track_emissions.py', 'compare_emissions.py', 'GreenCode']

    # Ensure the 'result' directory exists
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
    
    # Adjust the path for emissions.csv to be within the 'result' directory with suffix
    emissions_csv = os.path.join(result_dir, f'emissions_{suffix}.csv')

    # Check if the CSV file exists, if not, create it and write the header
    if not os.path.exists(emissions_data_csv):
        with open(emissions_data_csv, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([
                "Application name", "File Type", "Timestamp", "Emissions (gCO2eq)",
                "Duration", "emissions_rate", "CPU Power (KWh)", "GPU Power (KWh)", "RAM Power (KWh)",
                "CPU Energy (Wh)", "GPU Energy (KWh)", "RAM Energy (Wh)", "Energy Consumed (Wh)", "Test Results"
            ])

    tracker = EmissionsTracker()

    # Process files for each language
    process_files_by_type(base_dir, emissions_data_csv, result_dir, '.py', excluded_files, tracker, get_python_test_command)
    process_files_by_type(base_dir, emissions_data_csv, result_dir, '.java', excluded_files, tracker, get_java_test_command)
    process_files_by_type(base_dir, emissions_data_csv, result_dir, '.cpp', excluded_files, tracker, get_cpp_test_command)
    process_files_by_type(base_dir, emissions_data_csv, result_dir, '.cs', excluded_files, tracker, get_cs_test_command)

    print(f"Emissions data and test results written to {emissions_data_csv}")

# Call process_folder for 'before' and 'after' emissions data
process_folder(SOURCE_DIRECTORY, os.path.join(RESULT_DIR, 'main_before_emissions_data.csv'), RESULT_DIR, 'before-in-detail')
process_folder(GREEN_REFINED_DIRECTORY, os.path.join(RESULT_DIR, 'main_after_emissions_data.csv'), RESULT_DIR, 'after-in-detail')

def generate_pdf_report(merged_df, total_emissions_before, total_emissions_after, last_run_emissions, result_dir):
    pdf = FPDF()

    # Create first page with three eye-catching boxes
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)

    # Overall total before emissions box
    pdf.set_fill_color(255, 182, 193)  # Light red background for box
    pdf.cell(190, 20, "Overall Total Before Emissions: {:.2f} gCO2eq".format(total_emissions_before), 0, 1, 'C', True)

    # Overall total after emissions box
    pdf.set_fill_color(173, 216, 230)  # Light blue background for box
    pdf.cell(190, 20, "Overall Total After Emissions: {:.2f} gCO2eq".format(total_emissions_after), 0, 1, 'C', True)

    # Last run total emissions box
    # (You'll also need to pass the `last_run_emissions` value in the parameters)
    pdf.set_fill_color(144, 238, 144)  # Light green background for box
    pdf.cell(190, 20, "Last Run Emissions: {:.2f} gCO2eq".format(last_run_emissions), 0, 1, 'C', True)

    # Embedded and Non-Embedded filtering code here (same as before)...
    # Embedded code types
    embedded_types = ['.html', '.css', '.xml', '.php', '.ts']
    non_embedded_types = ['.py', '.java', '.cpp', '.rb']

    # Filter for embedded code
    pdf.add_page()
    plt.figure(figsize=(8, 6))
    embedded_df = merged_df[merged_df['File Type'].isin(embedded_types)]

    if embedded_df.empty:
        # If no embedded files, display a message
        plt.text(0.5, 0.5, 'No embedded code files found', horizontalalignment='center', verticalalignment='center',
                 fontsize=16, color='red', transform=plt.gca().transAxes)
        plt.axis('off')  # Hide the axes
    else:
        # Generate the bar plot for embedded code
        sns.barplot(x="File Type", y="Emissions (gCO2eq)_before", data=embedded_df, color='red', label='Before')
        sns.barplot(x="File Type", y="Emissions (gCO2eq)_after", data=embedded_df, color='green', label='After')
        plt.title('Embedded Code: Before vs After Emissions')
        plt.ylabel('Emissions (gCO2eq)')
        plt.xlabel('File Type')
        plt.legend()

    # Save the chart or message as an image
    embedded_chart_path = os.path.join(result_dir, 'embedded_emissions_chart.png')
    plt.savefig(embedded_chart_path)
    plt.close()
    pdf.image(embedded_chart_path, x=10, y=40, w=190)

    # Filter for non-embedded code
    pdf.add_page()
    plt.figure(figsize=(8, 6))
    non_embedded_df = merged_df[merged_df['File Type'].isin(non_embedded_types)]

    if non_embedded_df.empty:
        # If no non-embedded files, display a message
        plt.text(0.5, 0.5, 'No non-embedded code files found', horizontalalignment='center', verticalalignment='center',
                 fontsize=16, color='red', transform=plt.gca().transAxes)
        plt.axis('off')  # Hide the axes
    else:
        # Generate the bar plot for non-embedded code
        sns.barplot(x="File Type", y="Emissions (gCO2eq)_before", data=non_embedded_df, color='red', label='Before')
        sns.barplot(x="File Type", y="Emissions (gCO2eq)_after", data=non_embedded_df, color='green', label='After')
        plt.title('Non-Embedded Code: Before vs After Emissions')
        plt.ylabel('Emissions (gCO2eq)')
        plt.xlabel('File Type')
        plt.legend()

    # Save the chart or message as an image
    non_embedded_chart_path = os.path.join(result_dir, 'non_embedded_emissions_chart.png')
    plt.savefig(non_embedded_chart_path)
    plt.close()
    pdf.image(non_embedded_chart_path, x=10, y=40, w=190)
    
    # 3. Last Run Emissions Graph
    pdf.add_page()
    plt.figure(figsize=(8, 6))

    # Filter the merged_df for the last run data (based on the latest timestamp)
    merged_df['Timestamp_after'] = pd.to_datetime(merged_df['Timestamp_after'])  # Ensure the timestamp is datetime
    last_run_timestamp = merged_df['Timestamp_after'].max()  # Get the most recent timestamp
    last_run_data = merged_df[merged_df['Timestamp_after'] == last_run_timestamp]  # Filter for the last run

    # Calculate last run before and after emissions
    last_run_before = last_run_data['Emissions (gCO2eq)_before'].sum()
    last_run_after = last_run_data['Emissions (gCO2eq)_after'].sum()

    # Plot the bar chart for the last run emissions
    plt.bar(['Before'], [last_run_before], color='red', label='Before')
    plt.bar(['After'], [last_run_after], color='green', label='After')
    plt.title('Last Run Emissions: Before vs After')
    plt.ylabel('Emissions (gCO2eq)')
    plt.legend()

    # Save and add the last run emissions chart to the PDF
    last_run_chart_path = os.path.join(result_dir, 'last_run_emissions_chart.png')
    plt.savefig(last_run_chart_path)
    plt.close()
    pdf.image(last_run_chart_path, x=10, y=40, w=190)

   # Add the day-by-day transcript chart
    pdf.add_page()
    plt.figure(figsize=(8, 6))

    # Convert timestamps to dates and group by date to get total emissions per day
    merged_df['Date_before'] = pd.to_datetime(merged_df['Timestamp_before']).dt.date
    merged_df['Date_after'] = pd.to_datetime(merged_df['Timestamp_after']).dt.date

    # Aggregate emissions by date (sum or mean, based on your use case)
    emissions_before_daily = merged_df.groupby('Date_before')['Emissions (gCO2eq)_before'].sum().reset_index()
    emissions_after_daily = merged_df.groupby('Date_after')['Emissions (gCO2eq)_after'].sum().reset_index()

    # Ensure both datasets have the same date range (in case any dates are missing from one)
    common_dates = set(emissions_before_daily['Date_before']).intersection(set(emissions_after_daily['Date_after']))
    emissions_before_daily = emissions_before_daily[emissions_before_daily['Date_before'].isin(common_dates)]
    emissions_after_daily = emissions_after_daily[emissions_after_daily['Date_after'].isin(common_dates)]

    # Plot the emissions trends for 'Before' and 'After' by date
    sns.lineplot(x=emissions_before_daily["Date_before"], y=emissions_before_daily["Emissions (gCO2eq)_before"], label='Before', color='red')
    sns.lineplot(x=emissions_after_daily["Date_after"], y=emissions_after_daily["Emissions (gCO2eq)_after"], label='After', color='green')

    # Rotate x-axis labels and set a more spaced tick interval
    plt.xticks(rotation=45, ha='right')  # Rotate the x-axis labels to prevent overlap
    plt.gca().xaxis.set_major_locator(plt.MaxNLocator(6))  # Display fewer date ticks for clarity

    # Save and add the day-by-day chart to the PDF
    plt.title('Day-by-Day Emissions Trend')
    plt.ylabel('Emissions (gCO2eq)')
    plt.xlabel('Day')
    trend_chart_path = os.path.join(result_dir, 'emissions_trend_chart.png')
    plt.savefig(trend_chart_path)
    plt.close()

    pdf.image(trend_chart_path, x=10, y=40, w=190)


    # Save the final PDF report
    report_path = os.path.join(result_dir, 'emissions_report.pdf')
    pdf.output(report_path)
    print(f"PDF report generated and saved to {report_path}")

def compare_emissions():
    # Load environment variables again (if needed)
    load_dotenv(dotenv_path=env_path, verbose=True, override=True)

    # Remove the '.env' part to get the SOURCE_DIRECTORY
    result_source_dir = os.path.join(SOURCE_DIRECTORY, 'Result', 'main_before_emissions_data.csv')
    result_green_refined_directory = os.path.join(SOURCE_DIRECTORY, 'Result', 'main_after_emissions_data.csv')

    # Read CSV files
    emissions_df = pd.read_csv(result_source_dir)
    emissions_after_df = pd.read_csv(result_green_refined_directory)

    # Merge dataframes on common columns
    merged_df = emissions_df.merge(emissions_after_df, on=["Application name", "File Type"], suffixes=('_before', '_after'))

    # Calculate the difference in emissions and determine the result
    merged_df['final emission'] = merged_df['Emissions (gCO2eq)_before'] - merged_df['Emissions (gCO2eq)_after']
    merged_df['Result'] = merged_df['final emission'].apply(lambda x: 'Improved' if x > 0 else 'Need improvement')

    # Select and rename columns
    result_df = merged_df[["Application name", "File Type", "Timestamp_before", "Timestamp_after", "Emissions (gCO2eq)_before", "Emissions (gCO2eq)_after", "final emission", "Result"]]
    result_df.columns = ["Application name", "File Type", "Timestamp (Before)", "Timestamp (After)", "Before", "After", "Final Emission", "Result"]

    # Calculate the total emissions for the 'Before' and 'After'
    total_emissions_before = merged_df["Emissions (gCO2eq)_before"].sum()
    total_emissions_after = merged_df["Emissions (gCO2eq)_after"].sum()

    # Calculate the last run's emissions by finding the most recent timestamp
    merged_df = merged_df.sort_values(by='Timestamp_after', ascending=False)
    last_run_timestamp = merged_df['Timestamp_after'].iloc[0]
    last_run_data = merged_df[merged_df['Timestamp_after'] == last_run_timestamp]
    last_run_emissions = last_run_data['Emissions (gCO2eq)_after'].sum()

    # Create 'Result' folder if it doesn't exist
    if not os.path.exists(RESULT_DIR):
        os.makedirs(RESULT_DIR)

    # Write to new CSV file
    result_file_path = os.path.join(RESULT_DIR, "comparison_results.csv")
    result_df.to_csv(result_file_path, index=False)

    # Call the PDF generation function
    generate_pdf_report(merged_df, total_emissions_before, total_emissions_after, last_run_emissions, RESULT_DIR)

# Call the compare_emissions function
compare_emissions()
